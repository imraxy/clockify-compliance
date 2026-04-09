from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def compliance_to_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{payload.get('year')}-{payload.get('month'):02d}"
    header = Font(bold=True)
    ws.append(["S.N.", "Email", "Name", "Day columns →"])
    for c in range(1, 5):
        ws.cell(row=1, column=c).font = header
    year = int(payload["year"])
    month = int(payload["month"])
    from calendar import monthrange

    _, last = monthrange(year, month)
    day_cols = list(range(1, last + 1))
    for col, d in enumerate(day_cols, start=4):
        ws.cell(row=1, column=col, value=f"{year}-{month:02d}-{d:02d}")

    for idx, row in enumerate(payload.get("rows") or [], start=1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=row.get("email"))
        ws.cell(row=r, column=3, value=row.get("full_name"))
        days = row.get("days") or {}
        for col, d in enumerate(day_cols, start=4):
            cell = days.get(str(d)) or {}
            ws.cell(row=r, column=col, value=cell.get("status", ""))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
