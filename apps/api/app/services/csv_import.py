from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AttendanceDay, User
from app.services.merge_entries import resolve_user_by_email, upsert_time_entry


def _parse_dt(value: str) -> datetime:
    v = value.strip()
    if v.endswith("Z"):
        v = v[:-1] + "+00:00"
    return datetime.fromisoformat(v)


def _read_excel_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    """Read Excel file and return list of row dicts."""
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    if not ws:
        return []
    
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").lower().strip())
    
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row_dict = {}
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            value = cell.value
            if isinstance(value, datetime):
                row_dict[header] = value.isoformat()
            elif value is None:
                row_dict[header] = ""
            else:
                row_dict[header] = str(value).strip()
        rows.append(row_dict)
    
    wb.close()
    return rows


def import_time_entries_csv(db: Session, text: str) -> dict[str, Any]:
    """
    CSV columns: employee_email,start,end,description,project
    ISO-8601 datetimes for start/end.
    """
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"imported": 0, "errors": ["empty csv"]}
    errors: list[str] = []
    count = 0
    for i, row in enumerate(rows, start=2):
        try:
            email = (row.get("employee_email") or "").strip().lower()
            if not email:
                errors.append(f"row {i}: missing employee_email")
                continue
            user = resolve_user_by_email(db, email)
            if not user:
                errors.append(f"row {i}: unknown user {email}")
                continue
            start = _parse_dt(row.get("start") or "")
            end = _parse_dt(row.get("end") or "")
            desc = (row.get("description") or "").strip()
            proj = (row.get("project") or "").strip()
            upsert_time_entry(
                db,
                user_id=user.id,
                start=start,
                end=end,
                description=desc,
                project_name=proj,
                external_id=None,
                source="csv",
            )
            count += 1
        except Exception as e:  # noqa: BLE001 — surface row errors
            errors.append(f"row {i}: {e}")
    db.commit()
    return {"imported": count, "errors": errors}


def import_time_entries_excel(db: Session, file_bytes: bytes) -> dict[str, Any]:
    """Import from Excel with columns: employee_email, start, end, description, project"""
    rows = _read_excel_rows(file_bytes)
    if not rows:
        return {"imported": 0, "errors": ["empty file"]}
    
    errors: list[str] = []
    count = 0
    for i, row in enumerate(rows, start=2):
        try:
            email = (row.get("employee_email") or "").strip().lower()
            if not email:
                errors.append(f"row {i}: missing employee_email")
                continue
            user = resolve_user_by_email(db, email)
            if not user:
                errors.append(f"row {i}: unknown user {email}")
                continue
            start = _parse_dt(row.get("start") or "")
            end = _parse_dt(row.get("end") or "")
            desc = (row.get("description") or "").strip()
            proj = (row.get("project") or "").strip()
            upsert_time_entry(
                db,
                user_id=user.id,
                start=start,
                end=end,
                description=desc,
                project_name=proj,
                external_id=None,
                source="excel",
            )
            count += 1
        except Exception as e:
            errors.append(f"row {i}: {e}")
    db.commit()
    return {"imported": count, "errors": errors}


def import_attendance_csv(db: Session, text: str) -> dict[str, Any]:
    """CSV columns: employee_email, date, code"""
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    errors: list[str] = []
    count = 0
    for i, row in enumerate(rows, start=2):
        try:
            email = (row.get("employee_email") or "").strip().lower()
            day_s = (row.get("date") or "").strip()
            code = (row.get("code") or "").strip()
            if not email or not day_s or not code:
                errors.append(f"row {i}: missing fields")
                continue
            user = resolve_user_by_email(db, email)
            if not user:
                errors.append(f"row {i}: unknown user {email}")
                continue
            d = date.fromisoformat(day_s)
            existing = db.scalars(
                select(AttendanceDay).where(
                    AttendanceDay.user_id == user.id,
                    AttendanceDay.day == d,
                )
            ).first()
            if existing:
                existing.code = code
            else:
                db.add(AttendanceDay(user_id=user.id, day=d, code=code))
            count += 1
        except Exception as e:  # noqa: BLE001
            errors.append(f"row {i}: {e}")
    db.commit()
    return {"imported": count, "errors": errors}


def import_attendance_excel(db: Session, file_bytes: bytes) -> dict[str, Any]:
    """Import attendance from Attendance.xlsx format.
    
    Expected format (Sublimitysoft specific):
    - Row 1: Year with value 2026
    - Row 2: Month names (Jan, Feb, etc.)
    - Row 3: Headers - Name, Employee Number, Email Id, Phone, then date columns
    - Rows 4+: Employee data with attendance codes (P, WO, EL, WFH, H, etc.)
    
    Columns:
    - Col 1: Select (boolean)
    - Col 2: Name
    - Col 3: Employee Number
    - Col 4: Email Id
    - Col 5: Phone
    - Col 6: Empty
    - Col 7+: Date headers with datetime values, cells contain attendance codes
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    if not ws:
        return {"imported": 0, "errors": ["empty file"]}
    
    errors: list[str] = []
    count = 0
    
    # Find email column (column 4 - "Email Id")
    email_col = 4
    
    # Get date columns from row 3
    date_cols = {}
    for col_num in range(7, ws.max_column + 1):
        cell_val = ws.cell(row=3, column=col_num).value
        if isinstance(cell_val, datetime):
            date_cols[col_num] = cell_val.date()
    
    # Process each employee row (starting from row 4)
    for row_num in range(4, ws.max_row + 1):
        email = str(ws.cell(row=row_num, column=email_col).value or "").strip().lower()
        if not email:
            continue
        
        user = resolve_user_by_email(db, email)
        if not user:
            errors.append(f"row {row_num}: unknown user {email}")
            continue
        
        # Process each date column
        for col_num, d in date_cols.items():
            code = str(ws.cell(row=row_num, column=col_num).value or "").strip().upper()
            if not code or code in ["-", "", "NA", "N/A", "NONE"]:
                continue
            
            try:
                existing = db.scalars(
                    select(AttendanceDay).where(
                        AttendanceDay.user_id == user.id,
                        AttendanceDay.day == d,
                    )
                ).first()
                if existing:
                    existing.code = code
                else:
                    db.add(AttendanceDay(user_id=user.id, day=d, code=code))
                count += 1
            except Exception as e:
                errors.append(f"row {row_num} col {col_num}: {e}")
    
    db.commit()
    wb.close()
    return {"rows": count, "errors": errors}